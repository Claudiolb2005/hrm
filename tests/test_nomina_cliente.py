"""
Pruebas del motor Nomina Cliente (INCIDENCIAS -> Nomina Cliente).
Valida contra el par real SEM 26 de PRIME que el motor reproduce el entregable.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from motor.nomina_cliente import (  # noqa: E402
    build_master, build_isn_rates, build_rt_primas, generar, escribir_xlsx,
    NOM_HEADERS, O, _num, cesantia_patronal_rate, isr_retencion_semanal, UMA,
)

DATA = Path(__file__).resolve().parent.parent / "data" / "prime-nomina-cliente"
INC26 = DATA / "INCIDENCIAS_SEM26.xlsx"
NOM26 = DATA / "NominaCliente_SEM26.xlsx"
INC24 = DATA / "INCIDENCIAS_SEM24.xlsx"


@pytest.fixture(scope="module")
def comparacion():
    ref = [NOM26]
    master = build_master(ref)
    isn = build_isn_rates(ref)
    rt = build_rt_primas(ref)
    _, filas = generar(INC26, master, isn, rt)
    gen = {str(f[O["NO."]]).strip(): f for f in filas}
    wb = openpyxl.load_workbook(NOM26, read_only=True, data_only=True)
    real = [list(r) for r in wb.worksheets[0].iter_rows(values_only=True)][1:]
    wb.close()
    real = {str(r[1]).strip(): r for r in real if r[1] not in (None, "")}
    common = [k for k in gen if k in real]
    return gen, real, common


def _match(gen, real, common, header, tol):
    ci = NOM_HEADERS.index(header)
    ok = sum(1 for k in common if abs(_num(gen[k][ci]) - _num(real[k][ci])) <= tol)
    return ok / len(common)


def test_join_completo(comparacion):
    gen, real, common = comparacion
    assert len(common) >= 3100  # ~3131 empleados casan por ID


def test_columnas_directas_exactas(comparacion):
    gen, real, common = comparacion
    for h in ["SD", "SBC", "Sueldo", "Comisiones", "Prima dominical", "Despensa(Informativo)",
              "Préstamo Infonavit", "PENSION ALIMENTICIA", "Impuesto estatal"]:
        assert _match(gen, real, common, h, 0.011) >= 0.98, h


def test_estatutarias_desde_sbc(comparacion):
    gen, real, common = comparacion
    assert _match(gen, real, common, "I.M.S.S.", 1.0) >= 0.96
    assert _match(gen, real, common, "rcv (MONTO)", 1.0) >= 0.96
    assert _match(gen, real, common, "Infonavit empresa", 1.0) >= 0.98
    assert _match(gen, real, common, "Imss", 1.0) >= 0.96          # patronal (prima por RP)
    assert _match(gen, real, common, "RCV", 1.0) >= 0.96


def test_isr_y_neto(comparacion):
    gen, real, common = comparacion
    assert _match(gen, real, common, "I.S.R. (sp)", 1.0) >= 0.93
    assert _match(gen, real, common, "NETO SUELDO FISCAL", 1.0) >= 0.93


def test_precision_global(comparacion):
    gen, real, common = comparacion
    def isnum(h):
        return h not in ("NO. SEMANAL", "NO.", "Empleado", "RFC", "Puesto",
                         "Sucursal", "Registro patronal", "FECHA DE INGRESO")
    cols = [h for h in NOM_HEADERS if isnum(h)]
    prom = sum(_match(gen, real, common, h, 1.0) for h in cols) / len(cols)
    assert prom >= 0.99  # promedio de columnas dentro de $1


def test_cesantia_patronal_tramos():
    assert cesantia_patronal_rate(1.0 * UMA) == 0.03150
    assert cesantia_patronal_rate(2.8 * UMA) == 0.06026
    assert cesantia_patronal_rate(5.0 * UMA) == 0.07513


def test_isr_salario_minimo():
    # Asalariado de salario mínimo (~2206.68/sem) -> ISR ~29.42 tras subsidio 555.91
    isr, subs = isr_retencion_semanal(2206.68)
    assert abs(isr - 29.42) < 0.05
    assert subs == 0.0


def test_genera_y_escribe_sem24(tmp_path):
    ref = [NOM26]
    master, isn, rt = build_master(ref), build_isn_rates(ref), build_rt_primas(ref)
    sem, filas = generar(INC24, master, isn, rt)
    assert sem == 24
    assert len(filas) >= 3100
    out = tmp_path / "Nomina Cliente SEM 24.xlsx"
    escribir_xlsx(filas, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.title == "Base Pre nómina"
    assert ws.max_column == len(NOM_HEADERS)
    assert str(ws.cell(2, O["TOTAL PERCEPCIONES"] + 1).value).startswith("=SUM(")
    wb.close()
