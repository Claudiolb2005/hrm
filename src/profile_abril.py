"""Perfilado conciso de insumos ABRIL 2026 vs baseline Enero."""
import sys
from openpyxl import load_workbook

def profile(path, label):
    print(f"\n{'='*70}\n{label}: {path.split('/')[-1]}")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print("  ERROR abriendo:", e); return
    for ws in wb.worksheets:
        d = ws.max_row, ws.max_column
        print(f"  hoja '{ws.title}': {d[0]} filas x {d[1]} cols")
        # primeras 2 filas (encabezados probables), truncadas
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
            vals = [str(c)[:22] if c is not None else "" for c in row]
            rows.append(vals)
            if i >= 2: break
        # imprime encabezado no vacío más probable
        for ri, vals in enumerate(rows):
            nonempty = [v for v in vals if v]
            if nonempty:
                print(f"    fila{ri+1} ({len(nonempty)} no-vac): " + " | ".join(vals[:30]))
    wb.close()

base = "/Users/claudioleon/.openclaw/workspace/hrm/data/abril2026/"
profile(base+"acumulado_abril2026.xlsx", "ACUMULADO")
profile(base+"cfdi_abril2026.xlsx", "CFDI")
profile(base+"contabilidad_abril2026.xlsx", "CONTABILIDAD")
profile(base+"isn_abril2026.xlsx", "ISN")
