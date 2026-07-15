#!/usr/bin/env python3
"""Perfilador de insumos HRM (Enero 2026). Vuelca estructura real a stdout.
Sin dependencias del motor: solo openpyxl + pandas para inspección."""
import sys
from openpyxl import load_workbook
import pandas as pd

DATA = "/Users/claudioleon/.openclaw/workspace/hrm/data"


def sheets_overview(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append((ws.title, ws.max_row, ws.max_column))
    wb.close()
    return out


def dump_sheet_head(path, sheet, nrows=6, maxcols=200):
    """Lee las primeras nrows filas crudas (sin asumir encabezado)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row[:maxcols])
        if i + 1 >= nrows:
            break
    wb.close()
    return rows


def guess_header_row(rows):
    """Heurística simple: la fila con más celdas string no vacías."""
    best_i, best_score = 0, -1
    for i, r in enumerate(rows):
        score = sum(1 for c in r if isinstance(c, str) and c.strip())
        if score > best_score:
            best_i, best_score = i, score
    return best_i


def profile_file(label, filename, focus_sheets=None):
    path = f"{DATA}/{filename}"
    print("\n" + "=" * 90)
    print(f"### {label}  ({filename})")
    print("=" * 90)
    ov = sheets_overview(path)
    print(f"Hojas ({len(ov)}):")
    for t, r, c in ov:
        print(f"  - {t!r}: {r} filas x {c} cols")
    targets = focus_sheets or [ov[0][0]]
    for sheet in targets:
        if sheet not in [t for t, _, _ in ov]:
            print(f"  [!] hoja {sheet!r} no existe, salto")
            continue
        print(f"\n--- HOJA {sheet!r}: primeras filas ---")
        head = dump_sheet_head(path, sheet, nrows=5)
        hr = guess_header_row(head)
        print(f"  (fila de encabezado estimada: idx {hr})")
        header = head[hr]
        cols = [str(c).strip() if c is not None else "" for c in header]
        print(f"  columnas ({len(cols)}):")
        for j, name in enumerate(cols):
            if name:
                print(f"    [{j:>3}] {name}")
        # una fila de datos de muestra (la siguiente no vacía tras encabezado)
        for r in head[hr + 1:]:
            if any(c is not None and str(c).strip() for c in r):
                print("  muestra fila datos:")
                for j, (name, val) in enumerate(zip(cols, r)):
                    if name:
                        print(f"    [{j:>3}] {name} = {val!r}")
                break


if __name__ == "__main__":
    which = sys.argv[1] if len(sys.argv) > 1 else "all"
    if which in ("all", "acumulado"):
        profile_file("ACUMULADO", "acumulado_enero2026.xlsx")
    if which in ("all", "cfdi"):
        profile_file("CFDI (One Facture)", "cfdi_enero2026.xlsx")
    if which in ("all", "contab"):
        profile_file("CONTABILIDAD", "contabilidad_enero2026.xlsx")
    if which in ("all", "maestro"):
        profile_file("MAESTRO 01 ENERO", "maestro_enero2026.xlsx",
                     focus_sheets=["RESUMEN", "BASES"])
