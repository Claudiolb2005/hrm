"""Valida CFDI Abril (columnas por nombre) + conciliación R1/R6 vs acumulado."""
from motor.acumulado import parse_acumulado
from motor.cfdi import parse_cfdi, _resolver_columnas
from motor.concilia import conciliar_r1, conciliar_r6
from openpyxl import load_workbook

def m(c): return f"{c/100:,.2f}"

# comprobar que el resolver ubica bien las columnas de Abril
wb = load_workbook('data/abril2026/cfdi_abril2026.xlsx', read_only=True); ws = wb['Datos Cfdis']
hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True)); wb.close()
col = _resolver_columnas(hdr)
print("Columnas resueltas en Abril:", {k: col[k] for k in ('total','total_deducciones','total_otros_pagos','total_percepciones')})

acum = parse_acumulado('data/abril2026/acumulado_abril2026.xlsx')
print("ACUM neto_concil total:", m(acum.total.comprobacion_neto), "| RFCs:", len(acum.por_rfc))

for filtro in (None, '2026-04'):
    cf = parse_cfdi('data/abril2026/cfdi_abril2026.xlsx', mes_pago=filtro)
    print(f"\n--- CFDI mes_pago={filtro} ---")
    print(f"  cfdis {cf.meta['cfdis_contados']} | dup {cf.meta['duplicados_excluidos']} | canc {cf.meta['cancelados_excluidos']} | fuera_mes {cf.meta['fuera_de_mes']} | rfcs {cf.meta['rfcs_unicos']}")
    print(f"  neto {m(cf.total.neto)} (perc {m(cf.total.percepciones)} + otros {m(cf.total.otros_pagos)} - ded {m(cf.total.deducciones)})")
    r1 = conciliar_r1(acum.por_rfc, cf.por_rfc)
    r6 = conciliar_r6(acum.por_rfc, cf.por_rfc)
    print(f"  R1: acum {m(r1.acum_total)} vs cfdi {m(r1.cfdi_total)} | dif {m(r1.diferencia)} | emp_ok {r1.empleados_ok} emp_dif {r1.empleados_dif}")
    print(f"  R6: en_ambos {r6.en_ambos} | solo_acum {len(r6.solo_acumulado)} | solo_cfdi {len(r6.solo_cfdi)}")
