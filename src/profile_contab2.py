"""Composición de la dispersión (cuenta BBVA 006) por tipo de REFERENCE → R2."""
import re
from collections import defaultdict
from openpyxl import load_workbook

P = "data/abril2026/contabilidad_abril2026.xlsx"
wb = load_workbook(P, read_only=True, data_only=True); ws = wb["2489"]
rows = list(ws.iter_rows(values_only=True)); wb.close()
hdr = rows[0]
J = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
c_acc, c_desc, c_deb, c_cred, c_ref, c_date = (
    J["ACCOUNT"], J["ACCOUNT DESCRIPTION DETAIL"], J["DEBIT AMOUNT"],
    J["CREDIT AMOUNT"], J["REFERENCE"], J["TRANSACTION DATE"])
def num(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0
def clase(ref):
    r = (ref or "").upper()
    if "FINIQUITO" in r: return "FINIQUITO"
    if "PAGO DE NOMINA" in r or "PAGO NOMINA" in r: return "PAGO DE NOMINA"
    if "TRASPASO" in r: return "TRASPASO"
    if "SPEI" in r: return "SPEI"
    if "RECHAZO" in r or "DEVOL" in r: return "RECHAZO/DEVOL"
    return "OTRO: "+r[:22]

# descripciones de cuenta
desc = {}
for r in rows[1:]:
    a = str(r[c_acc]).strip() if r[c_acc] is not None else ""
    d = str(r[c_desc]).strip() if r[c_desc] is not None else ""
    if a and a not in desc and d: desc[a] = d
print("=== Descripción de cuentas ===")
for a,d in desc.items(): print(f"  {a}: {d}")

# composición por cuenta y clase de REFERENCE
for cuenta in ("10000-9000-006","22000-9000-000","10000-9000-001",""):
    comp = defaultdict(lambda:{"deb":0.0,"cred":0.0,"n":0})
    dmin=dmax=None
    for r in rows[1:]:
        a = str(r[c_acc]).strip() if r[c_acc] is not None else ""
        if a != cuenta: continue
        k = clase(r[c_ref])
        comp[k]["deb"]+=num(r[c_deb]); comp[k]["cred"]+=num(r[c_cred]); comp[k]["n"]+=1
        d=r[c_date]
        if d is not None:
            dmin=d if dmin is None or d<dmin else dmin
            dmax=d if dmax is None or d>dmax else dmax
    if not comp: continue
    print(f"\n=== Cuenta {cuenta or '(vacía)'} — {desc.get(cuenta,'?')} | fechas {str(dmin)[:10]}..{str(dmax)[:10]} ===")
    tot_d=tot_c=0
    for k,v in sorted(comp.items(), key=lambda x:-(x[1]['deb']+x[1]['cred'])):
        print(f"  {k:26} n={v['n']:>5}  D={v['deb']:>16,.2f}  C={v['cred']:>16,.2f}")
        tot_d+=v['deb']; tot_c+=v['cred']
    print(f"  {'TOTAL':26} {'':>7}  D={tot_d:>16,.2f}  C={tot_c:>16,.2f}")
