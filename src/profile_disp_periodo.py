"""¿Cuánta info de periodo trae la dispersión (cuenta 006)? Diseña Sec XI fino."""
import re
from collections import defaultdict
from openpyxl import load_workbook

P = "data/abril2026/contabilidad_abril2026.xlsx"
wb = load_workbook(P, read_only=True, data_only=True); ws = wb["2489"]
rows = list(ws.iter_rows(values_only=True)); wb.close()
J = {str(h).strip(): i for i, h in enumerate(rows[0]) if h is not None}
c_acc, c_cred, c_ref, c_date = J["ACCOUNT"], J["CREDIT AMOUNT"], J["REFERENCE"], J["TRANSACTION DATE"]
def num(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0

def periodo(ref):
    r = (ref or "").upper().replace(" ", "")
    m = re.search(r"NOMSEM0?(\d+)", r)
    if m: return f"SEM {int(m.group(1)):02d}"
    m = re.search(r"NOMCAT0?(\d+)", r) or re.search(r"CATORC0?(\d+)", r)
    if m: return f"CAT {int(m.group(1)):02d}"
    if "FINIQUITO" in r or "LIQUIDACION" in r: return "FINIQUITOS"
    if "PAGODENOMINA" in r: return "PAGO_GENERICO"
    return "OTRO"

# créditos del banco de nómina por periodo-derivado + por fecha
por_per = defaultdict(lambda: {"c": 0.0, "n": 0})
generico_por_fecha = defaultdict(lambda: {"c": 0.0, "n": 0})
for r in rows[1:]:
    if not r or str(r[c_acc]).strip() != "10000-9000-006": continue
    cred = num(r[c_cred])
    if not cred: continue
    per = periodo(r[c_ref])
    por_per[per]["c"] += cred; por_per[per]["n"] += 1
    if per == "PAGO_GENERICO":
        d = str(r[c_date])[:10]
        generico_por_fecha[d]["c"] += cred; generico_por_fecha[d]["n"] += 1

def m(x): return f"{x:,.2f}"
print("=== Créditos dispersión (006) por periodo derivado del REFERENCE ===")
for k, v in sorted(por_per.items(), key=lambda x: -x[1]["c"]):
    print(f"  {k:14} n={v['n']:>5}  {m(v['c']):>16}")
print("\n=== 'PAGO DE NOMINA' genérico desglosado por TRANSACTION DATE ===")
for d, v in sorted(generico_por_fecha.items()):
    print(f"  {d}  n={v['n']:>5}  {m(v['c']):>16}")
