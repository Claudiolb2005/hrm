"""Detalle fino ABRIL: BASES ISN2026, RESUMEN por estado, headers ACUMULADO, calendario."""
from openpyxl import load_workbook
base = "/Users/claudioleon/.openclaw/workspace/hrm/data/abril2026/"

def rows(path, sheet, maxr=200):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    out = []
    for r in ws.iter_rows(min_row=1, max_row=maxr, values_only=True):
        out.append(r)
    wb.close()
    return out

print("### ISN · hoja BASES (tasas 2026 por estado) ###")
for r in rows(base+"isn_abril2026.xlsx","BASES",40):
    if any(c is not None for c in r):
        print("  ", " | ".join(str(c) if c is not None else "" for c in r))

print("\n### ISN · hoja RESUMEN (base/tasa/impuesto por RP-estado = verdad Sec X) ###")
for r in rows(base+"isn_abril2026.xlsx","RESUMEN",40):
    vals=[c for c in r if c is not None]
    if len(vals)>=4:
        print("  ", " | ".join(str(c) if c is not None else "" for c in r))

print("\n### ACUMULADO · hoja 'Hoja1' (calendario periodos + total) ###")
for r in rows(base+"acumulado_abril2026.xlsx","Hoja1",20):
    if any(c is not None for c in r):
        print("  ", " | ".join(str(c) if c is not None else "" for c in r))

print("\n### ACUMULADO · 84 encabezados (fila 1) ###")
hdr = rows(base+"acumulado_abril2026.xlsx","Base Pre nómina",1)[0]
for i,h in enumerate(hdr):
    print(f"  [{i}] {h}")

print("\n### ACUMULADO · valores distintos en NO. PROCESO (periodos) ###")
wb=load_workbook(base+"acumulado_abril2026.xlsx",read_only=True,data_only=True)
ws=wb["Base Pre nómina"]
periodos={}
for r in ws.iter_rows(min_row=2,values_only=True):
    p=r[0]
    if p is not None: periodos[str(p)]=periodos.get(str(p),0)+1
wb.close()
for p,n in sorted(periodos.items()):
    print(f"  {p}: {n} filas")
