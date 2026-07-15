"""Perfila FEB 2026 (papel de trabajo): hoja de datos + ALTAS Y BAJAS + resumen."""
from openpyxl import load_workbook

P = "data/feb2026/acumulado_feb2026.xlsx"
wb = load_workbook(P, read_only=True, data_only=True)

print("=== Dimensiones de hojas ===")
cand = []
for ws in wb.worksheets:
    r, c = ws.max_row or 0, ws.max_column or 0
    print(f"  {ws.title:26} {r:>6} x {c}")
    if r > 1000 and c > 30:
        cand.append(ws.title)

# hoja de datos por empleado: encabezado con periodo/empleado/rfc
print(f"\n=== Candidatas a acumulado por empleado: {cand} ===")
for name in cand:
    ws = wb[name]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
        txt = [str(x)[:16] for x in row[:14] if x is not None]
        if any(isinstance(x, str) and x.upper() in
               ("RFC", "EMPLEADO", "NO. PROCESO", "NO.", "PUESTO") for x in row):
            print(f"  [{name}] fila{i}: " + " | ".join(str(x)[:16] if x is not None else "" for x in row[:14]))
            break

# ALTAS Y BAJAS completa (solo columnas A-E, el bloque principal)
print("\n=== ALTAS Y BAJAS (col A-E) ===")
ws = wb["ALTAS Y BAJAS"]
for r in ws.iter_rows(min_row=1, max_row=40, max_col=5, values_only=True):
    if any(c is not None for c in r):
        print("  " + " | ".join(str(c)[:20] if c is not None else "" for c in r))
wb.close()
