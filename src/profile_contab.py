"""Perfila CONTABILIDAD ABRIL para diseñar R2 (dispersión) / R4 (provisión) → Sec XI."""
from collections import defaultdict
from openpyxl import load_workbook

P = "data/abril2026/contabilidad_abril2026.xlsx"
wb = load_workbook(P, read_only=True, data_only=True)
ws = wb["2489"]
rows = list(ws.iter_rows(values_only=True))
wb.close()
hdr = rows[0]
idx = {str(h).strip(): j for j, h in enumerate(hdr) if h is not None}
def col(*names):
    for n in names:
        for k, j in idx.items():
            if k.upper() == n.upper():
                return j
    return None
c_acc   = col("ACCOUNT")
c_desc  = col("ACCOUNT DESCRIPTION DE","ACCOUNT DESCRIPTION")
c_deb   = col("DEBIT AMOUNT")
c_cred  = col("CREDIT AMOUNT")
c_ref   = col("REFERENCE")
c_date  = col("TRANSACTION DATE")
c_mes   = col("MES AL QUE CORRESPONDE")
c_type  = col("TYPE TRANSACTION")
print("Encabezados:", [str(h) for h in hdr])
print(f"idx: account={c_acc} desc={c_desc} debit={c_deb} credit={c_cred} ref={c_ref} date={c_date} mes={c_mes}")
print(f"filas datos: {len(rows)-1}")

def num(v):
    try: return float(v)
    except (TypeError, ValueError): return 0.0

# agregado por cuenta (código + descripción)
agg = defaultdict(lambda: {"deb":0.0,"cred":0.0,"n":0})
for r in rows[1:]:
    if not r or c_acc is None: continue
    acc = str(r[c_acc]).strip() if r[c_acc] is not None else ""
    desc = str(r[c_desc]).strip() if c_desc is not None and r[c_desc] is not None else ""
    k = (acc, desc[:34])
    agg[k]["deb"]  += num(r[c_deb])  if c_deb  is not None else 0
    agg[k]["cred"] += num(r[c_cred]) if c_cred is not None else 0
    agg[k]["n"]    += 1

print("\n=== Cuentas (código | descripción | n | Σdébito | Σcrédito) ===")
for (acc,desc),v in sorted(agg.items(), key=lambda x:-(x[1]['deb']+x[1]['cred'])):
    print(f"  {acc:20} {desc:34} n={v['n']:>5}  D={v['deb']:>16,.2f}  C={v['cred']:>16,.2f}")

# patrones de REFERENCE (para detectar periodo) y MES
print("\n=== Muestra REFERENCE (20) ===")
refs = [str(r[c_ref]) for r in rows[1:] if c_ref is not None and r[c_ref] is not None][:20]
for x in refs: print("   ", x[:60])
print("\n=== Valores MES AL QUE CORRESPONDE ===")
meses = defaultdict(int)
for r in rows[1:]:
    if c_mes is not None and r[c_mes] is not None:
        meses[str(r[c_mes]).strip()] += 1
for k,n in sorted(meses.items()): print(f"   {k!r}: {n}")
print("\n=== Valores TYPE TRANSACTION ===")
tt=defaultdict(int)
for r in rows[1:]:
    if c_type is not None and r[c_type] is not None: tt[str(r[c_type]).strip()]+=1
for k,n in sorted(tt.items()): print(f"   {k!r}: {n}")
