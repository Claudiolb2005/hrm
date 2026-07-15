"""Valida Sec X (ISN por estado): Abril vs su RESUMEN; Enero con tasas 2026."""
from openpyxl import load_workbook
from motor.maestro import parse_resumen, parse_bases
from motor.secciones import seccion_x_isn

def m(c): return f"{c/100:,.2f}"

ABRIL = "data/abril2026/isn_abril2026.xlsx"
ENERO = "data/maestro_enero2026.xlsx"

print("========== ABRIL — Sec X vs RESUMEN (ground truth) ==========")
res_a = parse_resumen(ABRIL)
bases_a = parse_bases(ABRIL)
secx_a = seccion_x_isn(res_a, bases_a)
print(f"estados: {len(secx_a['filas'])} | BASES con tasa>0: {sum(1 for v in bases_a.values() if v['tasa']>0)}/{len(bases_a)}")
print(f"total base            : {m(secx_a['total_base'])}   (GT 126,983,860.52)")
print(f"total impuesto RESUMEN : {m(secx_a['total_impuesto'])}   (GT 4,195,743.80)")
print(f"total impuesto BASES26 : {m(secx_a['total_impuesto_bases'])}")
print(f"discrepancias tasa (RESUMEN vs BASES): {len(secx_a['discrepancias_tasa'])}")
print(f"renglones que descuadran (base*tasa!=imp): {len(secx_a['renglones_descuadran'])}")

print("\n========== ENERO — sheets del maestro ==========")
wb = load_workbook(ENERO, read_only=True); print("hojas:", wb.sheetnames); wb.close()

print("\n========== ENERO — Sec X con SU BASES vs con BASES 2026 (Abril) ==========")
try:
    res_e = parse_resumen(ENERO)
    bases_e = parse_bases(ENERO)
    secx_e_old = seccion_x_isn(res_e, bases_e)          # tasas de enero (posibles viejas)
    secx_e_new = seccion_x_isn(res_e, bases_a)          # tasas 2026 autoritativas
    print(f"estados: {len(res_e)} | base total: {m(secx_e_old['total_base'])}")
    print(f"ISN con tasas ENERO (BASES propias): {m(secx_e_old['total_impuesto_bases'])}")
    print(f"ISN con tasas 2026 (BASES abril)   : {m(secx_e_new['total_impuesto_bases'])}   (informe entregado: 4,558,094.94)")
    print(f"ISN del RESUMEN de enero            : {m(secx_e_old['total_impuesto'])}")
    # estados donde la tasa 2026 difiere de la de enero
    difs = []
    for r in res_e:
        te = bases_e.get(r['rp'],{}).get('tasa')
        ta = bases_a.get(r['rp'],{}).get('tasa')
        if te is not None and ta is not None and abs(te-ta) > 1e-9:
            difs.append((r['entidad'], te, ta, r['base_isn']))
    print(f"\nEstados con tasa distinta enero->2026 ({len(difs)}):")
    for ent, te, ta, base in sorted(difs, key=lambda x:-x[3]):
        delta = round(base*(ta-te))
        print(f"  {ent:20} {te:.4f} -> {ta:.4f}  base {m(base):>15}  ΔISN {m(delta):>12}")
    print(f"  Σ ΔISN = {m(sum(round(b*(ta-te)) for _,te,ta,b in difs))}")
except Exception as e:
    import traceback; traceback.print_exc()
