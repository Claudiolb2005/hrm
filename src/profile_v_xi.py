"""Datos para Sec V (catálogo puesto→familia) y Sec XI fino (Hoja1 provisión)."""
from collections import defaultdict
from openpyxl import load_workbook
from motor.acumulado import parse_acumulado

AC = "data/abril2026/acumulado_abril2026.xlsx"

# ---- Sec V: puestos crudos por tipo (sem/cat/esp) con importe ----
acum = parse_acumulado(AC)
# reconstruye puesto->tipo->monto desde el detalle (por_puesto ya trae sem/cat/esp)
sem = [(p, g["sem"]) for p, g in acum.por_puesto.items() if g["sem"]]
cat = [(p, g["cat"]) for p, g in acum.por_puesto.items() if g["cat"]]
def dump(titulo, items):
    items = sorted(items, key=lambda x: -x[1])
    print(f"\n=== {titulo}: {len(items)} puestos distintos | total {sum(v for _,v in items)/100:,.2f} ===")
    for p, v in items[:40]:
        print(f"  {v/100:>14,.2f}  {p}")
dump("SEMANAL (puestos crudos)", sem)
dump("CATORCENAL (puestos crudos)", cat)

# ---- Sec XI: Hoja1 (provisión por periodo) ----
print("\n\n########## Hoja1 del acumulado (provisión/calendario) ##########")
wb = load_workbook(AC, read_only=True, data_only=True)
ws = wb["Hoja1"]
for r in ws.iter_rows(values_only=True):
    if any(c is not None for c in r):
        print("  " + " | ".join(str(c) if c is not None else "" for c in r))
wb.close()
