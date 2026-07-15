"""Parser del ACUMULADO (insumo 1) — spec §5.3.3.

Determinista, en centavos enteros. Produce:
- agregados por periodo (base de secciones IV/VII/VIII/IX)
- totales del mes
- hallazgos (control de fila percepciones - deducciones = neto)

No toca IA, no adivina columnas requeridas en silencio (§5.3.2).
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from openpyxl import load_workbook

from .normaliza import (
    normalizar_clave, a_centavos, normalizar_rfc, es_fila_total, CoercionError,
)
from .conceptos_prime import (
    PLANTILLA_ACUMULADO, FAMILIAS_ACUMULADO, REQUERIDAS,
)

TOLERANCIA_FILA = 1  # ±1 centavo en el control de fila (§5.6)


@dataclass
class AgregadoPeriodo:
    empleados: int = 0
    percepciones: int = 0
    deducciones: int = 0
    despensa_informativo: int = 0
    neto: int = 0                 # NETO SUELDO FISCAL (perc - despensa - ded)
    comprobacion_neto: int = 0    # perc - ded (neto de conciliación, incluye despensa)
    isr: int = 0
    imss_obrero: int = 0
    rcv_obrero: int = 0
    infonavit: int = 0
    fonacot: int = 0
    imss_patronal: int = 0
    rcv_patronal: int = 0
    infonavit_patronal: int = 0
    isn_informativo: int = 0
    prevision_social: int = 0
    rfcs: set = field(default_factory=set)


@dataclass
class ResultadoAcumulado:
    periodos: dict          # periodo -> AgregadoPeriodo
    total: AgregadoPeriodo
    hallazgos: list
    meta: dict
    por_rfc: dict = None    # rfc -> {neto_concil, neto_fiscal, filas, periodos:set}
    por_rp: dict = None     # registro patronal -> {neto_concil, imss/rcv obrero+patronal, rfcs}
    por_puesto: dict = None # puesto -> {sem, cat, esp, filas}


def _detectar_fila_encabezado(filas, variantes_norm):
    """Devuelve el índice de la fila (0-9) con más celdas que mapean a canónicas."""
    mejor_i, mejor = 0, -1
    for i, fila in enumerate(filas[:10]):
        score = 0
        for c in fila:
            if isinstance(c, str) and normalizar_clave(c) in variantes_norm:
                score += 1
        if score > mejor:
            mejor_i, mejor = i, score
    return mejor_i, mejor


def _mapear_columnas(header, plantilla):
    """canónico -> índice de columna, por clave normalizada. Reporta faltantes."""
    norm_to_idx = {}
    for j, c in enumerate(header):
        if isinstance(c, str) and c.strip():
            k = normalizar_clave(c)
            norm_to_idx.setdefault(k, j)  # primera ocurrencia gana
    mapa = {}
    for canon, cabecera in plantilla.items():
        # el valor puede ser una cadena o una lista de variantes equivalentes
        variantes = cabecera if isinstance(cabecera, (list, tuple)) else [cabecera]
        for cab in variantes:
            idx = norm_to_idx.get(normalizar_clave(cab))
            if idx is not None:
                mapa[canon] = idx
                break
    faltantes = [c for c in REQUERIDAS if c not in mapa]
    return mapa, faltantes, norm_to_idx


def _mapear_familias(norm_to_idx):
    """familia -> lista de índices presentes."""
    fam = {}
    for nombre, cabeceras in FAMILIAS_ACUMULADO.items():
        idxs = []
        for cab in cabeceras:
            j = norm_to_idx.get(normalizar_clave(cab))
            if j is not None:
                idxs.append(j)
        fam[nombre] = idxs
    return fam


def _suma_idxs(fila, idxs):
    total = 0
    for j in idxs:
        if j < len(fila):
            total += a_centavos(fila[j])
    return total


def parse_acumulado(path, empresa_id="PRIME", plantilla=None):
    plantilla = plantilla or PLANTILLA_ACUMULADO
    variantes_norm = set()
    for v in plantilla.values():
        for vv in (v if isinstance(v, (list, tuple)) else [v]):
            variantes_norm.add(normalizar_clave(vv))
    for cabs in FAMILIAS_ACUMULADO.values():
        variantes_norm |= {normalizar_clave(c) for c in cabs}

    wb = load_workbook(path, read_only=True, data_only=True)
    # seleccionar hoja: la de más filas x columnas (ACUMULADO, §5.3.1)
    ws = max(wb.worksheets, key=lambda s: (s.max_row or 0) * (s.max_column or 0))
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    h, score = _detectar_fila_encabezado(filas, variantes_norm)
    if score <= 0:
        raise CoercionError("no se encontró fila de encabezado en el ACUMULADO")
    header = filas[h]
    mapa, faltantes, norm_to_idx = _mapear_columnas(header, plantilla)
    if faltantes:
        raise CoercionError(f"columnas requeridas faltantes: {faltantes}")
    fam = _mapear_familias(norm_to_idx)

    c = mapa  # alias
    periodos: dict = defaultdict(AgregadoPeriodo)
    total = AgregadoPeriodo()
    por_rfc: dict = defaultdict(lambda: {"neto_concil": 0, "neto_fiscal": 0,
                                          "filas": 0, "periodos": set()})
    por_rp: dict = defaultdict(lambda: {"neto_concil": 0, "imss_obrero": 0,
                                        "rcv_obrero": 0, "imss_patronal": 0,
                                        "rcv_patronal": 0, "infonavit_patronal": 0,
                                        "filas": 0, "rfcs": set()})
    por_puesto: dict = defaultdict(lambda: {"sem": 0, "cat": 0, "esp": 0, "filas": 0})
    hallazgos = []
    filas_datos = 0
    filas_skip = 0

    for r_i in range(h + 1, len(filas)):
        fila = filas[r_i]
        if fila is None:
            continue
        # fila totalmente vacía
        if not any(v is not None and str(v).strip() for v in fila):
            filas_skip += 1
            continue
        periodo_raw = fila[c["periodo"]] if c["periodo"] < len(fila) else None
        periodo = normalizar_clave(periodo_raw)
        # subtotal/total o llave vacía
        if not periodo or es_fila_total(periodo):
            filas_skip += 1
            continue

        try:
            perc = a_centavos(fila[c["percepciones"]])
            ded = a_centavos(fila[c["deducciones"]])
            neto = a_centavos(fila[c["neto"]])
            desp = a_centavos(fila[c["despensa_informativo"]]) if "despensa_informativo" in c else 0
        except CoercionError as e:
            hallazgos.append({"tipo": "coercion", "fila": r_i + 1,
                              "periodo": periodo, "detalle": str(e)})
            continue

        # control de fila (§5.3.3): percepciones - despensa(informativo) - deducciones
        # = NETO SUELDO FISCAL (±1 centavo). La despensa está en percepciones pero es
        # informativa y no llega al neto fiscal.
        comprob_neto = perc - ded            # neto de conciliación (incluye despensa)
        if abs((perc - desp - ded) - neto) > TOLERANCIA_FILA:
            rfc_raw = fila[c["rfc"]] if c["rfc"] < len(fila) else ""
            hallazgos.append({
                "tipo": "control_fila", "fila": r_i + 1, "periodo": periodo,
                "rfc": str(rfc_raw), "perc": perc, "ded": ded, "desp": desp,
                "neto": neto, "descuadre": (perc - desp - ded) - neto,
            })

        rfc, _ = normalizar_rfc(fila[c["rfc"]] if c["rfc"] < len(fila) else "")

        def val(canon):
            j = c.get(canon)
            return a_centavos(fila[j]) if (j is not None and j < len(fila)) else 0

        # valores de fila calculados UNA vez (se reparten a todos los agregados)
        rv = {
            "isr": _suma_idxs(fila, fam["isr"]),
            "infonavit": _suma_idxs(fila, fam["infonavit"]),
            "fonacot": _suma_idxs(fila, fam["fonacot"]),
            "imss_obrero": val("imss_obrero"), "rcv_obrero": val("rcv_obrero"),
            "imss_patronal": val("imss_patronal"), "rcv_patronal": val("rcv_patronal"),
            "infonavit_patronal": val("infonavit_patronal"),
            "isn_informativo": val("isn_informativo"),
            "prevision_social": val("prevision_social"),
        }

        for bucket in (periodos[periodo], total):
            bucket.empleados += 1
            bucket.percepciones += perc
            bucket.deducciones += ded
            bucket.despensa_informativo += desp
            bucket.neto += neto
            bucket.comprobacion_neto += comprob_neto
            bucket.isr += rv["isr"]
            bucket.infonavit += rv["infonavit"]
            bucket.fonacot += rv["fonacot"]
            bucket.imss_obrero += rv["imss_obrero"]
            bucket.rcv_obrero += rv["rcv_obrero"]
            bucket.imss_patronal += rv["imss_patronal"]
            bucket.rcv_patronal += rv["rcv_patronal"]
            bucket.infonavit_patronal += rv["infonavit_patronal"]
            bucket.isn_informativo += rv["isn_informativo"]
            bucket.prevision_social += rv["prevision_social"]
            if rfc:
                bucket.rfcs.add(rfc)
        if rfc:
            g = por_rfc[rfc]
            g["neto_concil"] += comprob_neto
            g["neto_fiscal"] += neto
            g["filas"] += 1
            g["periodos"].add(periodo)

        # por REGISTRO PATRONAL (Sec VI vía estado, Sec IX IMSS/RCV)
        rp = str(fila[c["registro_patronal"]]).strip() if c.get("registro_patronal") is not None \
            and c["registro_patronal"] < len(fila) and fila[c["registro_patronal"]] is not None else ""
        if rp:
            gr = por_rp[rp]
            gr["neto_concil"] += comprob_neto
            gr["imss_obrero"] += rv["imss_obrero"]
            gr["rcv_obrero"] += rv["rcv_obrero"]
            gr["imss_patronal"] += rv["imss_patronal"]
            gr["rcv_patronal"] += rv["rcv_patronal"]
            gr["infonavit_patronal"] += rv["infonavit_patronal"]
            gr["filas"] += 1
            if rfc:
                gr["rfcs"].add(rfc)

        # por PUESTO, separando SEM / CAT / especiales (Sec V)
        puesto = str(fila[c["puesto"]]).strip() if c.get("puesto") is not None \
            and c["puesto"] < len(fila) and fila[c["puesto"]] is not None else "(SIN PUESTO)"
        tipo = "sem" if periodo.startswith("SEM") else ("cat" if periodo.startswith("CAT") else "esp")
        gp = por_puesto[puesto]
        gp[tipo] += comprob_neto
        gp["filas"] += 1

        filas_datos += 1

    meta = {
        "archivo": path,
        "empresa_id": empresa_id,
        "hoja": ws.title,
        "fila_encabezado": h,
        "columnas_detectadas": len([x for x in header if x]),
        "filas_datos": filas_datos,
        "filas_omitidas": filas_skip,
        "familias": {k: len(v) for k, v in fam.items()},
        "empleados_unicos_mes": len(total.rfcs),
    }
    return ResultadoAcumulado(dict(periodos), total, hallazgos, meta,
                              dict(por_rfc), dict(por_rp), dict(por_puesto))
