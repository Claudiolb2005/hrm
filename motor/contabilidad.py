"""Parser de CONTABILIDAD (insumo 4) — libro mayor de pólizas. Spec §5.3.6.

Alimenta R2 (dispersión) y R4 (provisión) → Sección XI "Provisión Global
(contabilidad) vs Nómina", presente en TODOS los informes.

Doble partida (PRIME): pagar nómina DEBITA el PASIVO GLOBAL NOMINA (22000-9000)
y ACREDITA el banco de nómina BBVA 0117452489 (10000-9000-006). El fondeo entra
a ese banco como TRASPASO (89.4M desde la cuenta de uso general) — es el MISMO
dinero que luego sale, así que NO se cuenta como dispersión (regla v6 §5.7:
"pago de nómina" y "traspaso BBVA" = mismo dinero, contar una vez).

Determinista, centavos enteros. La IA no interviene.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from openpyxl import load_workbook

from .normaliza import a_centavos, normalizar_clave

# --- plantilla PRIME (catálogo de cuentas). Multi-cliente: otra empresa la suya ---
CUENTA_BANCO_NOMINA = "10000-9000-006"   # BBVA 0117452489 — dispersión de nómina
CUENTA_PASIVO_NOMINA = "22000-9000"       # PASIVO GLOBAL NOMINA (prefijo)
HOJA_CONTAB = "2489"


def _clase_ref(ref):
    """Clasifica una póliza por su REFERENCE (regla de dispersión v6)."""
    r = normalizar_clave(ref).replace(" ", "")
    if "FINIQUITO" in r or "LIQUIDACION" in r:
        return "finiquito"
    if "TRASPASO" in r:
        return "traspaso"
    if "RECHAZO" in r or "DEVOLUCION" in r or "DEVOL" in r:
        return "rechazo"
    if "PAGONOM" in r or "PAGODENOMINA" in r or "NOMSEM" in r or "NOMCAT" in r:
        return "pago_nomina"
    if "SPEI" in r:
        return "spei"
    return "otro"


@dataclass
class ResultadoContab:
    dispersado: int                 # Σ créditos del banco de nómina (salida a empleados)
    fondeo_traspaso: int            # Σ débitos de traspaso (fondeo entrante; mismo dinero)
    por_clase: dict = field(default_factory=dict)   # clase -> centavos (composición dispersión)
    pasivo_debito: int = 0          # liquidación del pasivo (pagos)
    pasivo_credito: int = 0         # provisión del pasivo
    hallazgos: list = field(default_factory=list)
    meta: dict = field(default_factory=dict)


def _cols(header):
    idx = {}
    for j, h in enumerate(header):
        if isinstance(h, str) and h.strip():
            idx.setdefault(normalizar_clave(h), j)
    def c(*nombres):
        for n in nombres:
            j = idx.get(normalizar_clave(n))
            if j is not None:
                return j
        return None
    return {
        "account": c("ACCOUNT"), "ref": c("REFERENCE"),
        "debito": c("DEBIT AMOUNT"), "credito": c("CREDIT AMOUNT"),
        "fecha": c("TRANSACTION DATE"), "desc": c("ACCOUNT DESCRIPTION DETAIL"),
    }


def parse_contabilidad(path, cuenta_banco=CUENTA_BANCO_NOMINA,
                       cuenta_pasivo=CUENTA_PASIVO_NOMINA, hoja=HOJA_CONTAB):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[hoja] if hoja in wb.sheetnames else \
        max(wb.worksheets, key=lambda s: (s.max_row or 0) * (s.max_column or 0))
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    col = _cols(filas[0])
    if col["account"] is None or col["credito"] is None:
        raise ValueError("CONTABILIDAD: faltan columnas ACCOUNT/CREDIT")

    dispersado = 0
    fondeo = 0
    por_clase = defaultdict(int)
    pas_d = pas_c = 0
    n_banco = 0
    fmin = fmax = None

    for r in filas[1:]:
        if not r or r[col["account"]] is None:
            continue
        acc = str(r[col["account"]]).strip()
        deb = a_centavos(r[col["debito"]]) if col["debito"] is not None and r[col["debito"]] is not None else 0
        cred = a_centavos(r[col["credito"]]) if r[col["credito"]] is not None else 0
        clase = _clase_ref(r[col["ref"]] if col["ref"] is not None else "")

        if acc == cuenta_banco:
            n_banco += 1
            # créditos = salida de dinero (dispersión a empleados)
            if cred:
                if clase == "traspaso":
                    fondeo += cred          # (raro como crédito, pero se aísla)
                else:
                    dispersado += cred
                    por_clase[clase] += cred
            # débitos = entrada de dinero (fondeo por traspaso)
            if deb and clase == "traspaso":
                fondeo += deb
            f = r[col["fecha"]]
            if f is not None:
                fmin = f if fmin is None or f < fmin else fmin
                fmax = f if fmax is None or f > fmax else fmax
        elif acc.startswith(cuenta_pasivo):
            pas_d += deb
            pas_c += cred

    meta = {
        "archivo": path, "hoja": ws.title, "cuenta_banco": cuenta_banco,
        "cuenta_pasivo": cuenta_pasivo, "movs_banco": n_banco,
        "fecha_min": str(fmin)[:10] if fmin else None,
        "fecha_max": str(fmax)[:10] if fmax else None,
    }
    return ResultadoContab(dispersado, fondeo, dict(por_clase), pas_d, pas_c, [], meta)


@dataclass
class ProvisionPeriodos:
    periodos: list = field(default_factory=list)     # [(nombre, provisión_centavos)]
    especiales: list = field(default_factory=list)   # [(nombre, pago_centavos)] BONOS/FINIQUITOS
    acumulado: int = 0        # Σ provisión de los periodos regulares
    pago: int = 0             # Σ dispersado de esos periodos
    dif: int = 0              # acumulado − pago (redondeos)
    meta: dict = field(default_factory=dict)


def parse_provision_periodos(path_acumulado, hoja="Hoja1"):
    """Lee la hoja de PROVISIÓN por periodo (calendario) que acompaña al ACUMULADO
    (PERIODO | rango fechas | fecha pago | importe | etiqueta) → base de la
    tabla de la Sección XI. Reproduce el papel de trabajo: por periodo el importe
    provisionado, y los controles ACUMULADO/PAGO/DIF + líneas especiales (BONOS,
    FINIQUITOS) marcadas como PAGO."""
    wb = load_workbook(path_acumulado, read_only=True, data_only=True)
    ws = wb[hoja] if hoja in wb.sheetnames else None
    if ws is None:
        for s in wb.worksheets:                       # fallback: hoja chica con periodos
            head = [str(r[0]) for r in s.iter_rows(min_row=1, max_row=4, values_only=True) if r and r[0]]
            if any("SEMANA" in h.upper() or "CATORCENA" in h.upper() for h in head):
                ws = s
                break
    filas = list(ws.iter_rows(values_only=True)) if ws else []
    wb.close()

    periodos, especiales = [], []
    acumulado = pago = dif = 0
    for r in filas:
        if not r:
            continue
        nombre = str(r[0]).strip() if r[0] is not None else ""
        importe = None
        for v in r[1:]:                               # primer numérico tras la etiqueta
            if isinstance(v, (int, float)):
                importe = a_centavos(v)
                break
        etiqueta = ""
        for v in r:
            if isinstance(v, str) and normalizar_clave(v) in ("ACUMULADO", "PAGO", "DIF"):
                etiqueta = normalizar_clave(v)
        if importe is None:
            continue
        nu = nombre.upper()
        if etiqueta == "ACUMULADO":
            acumulado = importe
        elif etiqueta == "DIF":
            dif = importe
        elif etiqueta == "PAGO" and nombre:
            especiales.append((nombre, importe))       # BONOS / FINIQUITOS
        elif etiqueta == "PAGO":
            pago = importe
        elif nu.startswith(("SEMANA", "CATORCENA")):
            periodos.append((nombre, importe))
    meta = {"archivo": path_acumulado, "hoja": ws.title if ws else None,
            "n_periodos": len(periodos)}
    return ProvisionPeriodos(periodos, especiales, acumulado, pago, dif, meta)


def parse_resumen_conciliacion(path_acumulado, hoja="resumen"):
    """Lee la hoja de conciliación por periodo del papel de trabajo (layout PRIME):
    ACUMULADO (nómina con despensa) | CFDI | PAGOS (dispersión) por periodo. Es la
    fuente de la columna VARIACIÓN de la Sección XI del mes vivo (la Hoja1/provisión
    trae el ciclo previo). → dict periodo -> {acumulado, cfdi, pagos} en centavos."""
    wb = load_workbook(path_acumulado, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        wb.close()
        return {}
    filas = list(wb[hoja].iter_rows(values_only=True))
    wb.close()
    # localiza el encabezado del pivot DETALLADO: col0='Etiquetas de fila' y en la
    # col 7 arranca el bloque PAGOS ('PERIODO'). Así se ignora el pivot simple de arriba.
    h = None
    for i, r in enumerate(filas):
        if (r and len(r) > 8 and isinstance(r[0], str)
                and normalizar_clave(r[0]) == "ETIQUETAS DE FILA"
                and isinstance(r[7], str) and normalizar_clave(r[7]) == "PERIODO"):
            h = i
            break
    if h is None:
        return {}
    out = {}
    for r in filas[h + 1:]:
        if not r or r[0] is None:
            continue
        periodo = str(r[0]).strip()
        if normalizar_clave(periodo) in ("", "TOTAL GENERAL", "EN BLANCO"):
            continue
        def cel(j):
            return a_centavos(r[j]) if j < len(r) and isinstance(r[j], (int, float)) else 0
        out[periodo] = {"acumulado": cel(1), "cfdi": cel(4), "pagos": cel(8)}
    return out
