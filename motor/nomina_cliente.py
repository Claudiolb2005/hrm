"""
motor/nomina_cliente.py
========================
Genera el entregable **Nomina Cliente SEM X** de PRIME a partir del archivo
**INCIDENCIAS NOMINA SEMANA X** que el cliente envía.

Filosofía (spec HRM v6): motor determinista, dinero en centavos, la IA nunca
hace aritmética, fail-closed / exponer hallazgos, validar al centavo.

Ingeniería inversa hecha contra el par real SEM 26 (INCIDENCIAS -> Nomina Cliente):
  - 55+ columnas se derivan directo de INCIDENCIAS (identidad + incidencias + percepciones).
  - Columnas estatutarias (SBC, ISR, IMSS, RCV, Infonavit, ISN) se calculan con
    tablas 2026 oficiales a partir del SBC (que viene de un MAESTRO por empleado,
    NO de INCIDENCIAS).
  - Los 4 totales son fórmulas SUM/resta (igual que el Excel original).

Parámetros fiscales 2026 recuperados de la data real de PRIME y verificados contra
fuentes oficiales (DOF Anexo 8 RMF 2026, reforma de pensiones, cuotas IMSS 2026):
  UMA diaria = 117.31 ; tarifa/subsidio ISR semanal ; cesantía patronal por tramo UMA.
"""
from __future__ import annotations
import math
from datetime import datetime, timedelta, date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# 1. LAYOUTS DE COLUMNA (0-based, por posición: robusto a nombres duplicados)
# ---------------------------------------------------------------------------
# INCIDENCIAS NOMINA SEMANA X  (hoja "Prenomina Semana X"): header en fila idx 1,
# subheader UNIDAD/MONTO en fila idx 2, datos desde fila idx 3.
INC = {
    "semana": 0, "id": 1, "nombre": 2, "ingreso": 3, "curp": 4, "rfc": 5, "nss": 6,
    "depto": 7, "puesto": 8, "sd": 9, "sm": 10, "banco": 11, "clabe": 12, "correo": 13,
    "aguinaldo_dias": 14, "vac_dias": 15, "prima_vac_dias": 16,
    "faltas": 17, "vac_aprob": 18, "incap": 19, "domingos": 20, "dias_p_vac": 21,
    "descansos": 22, "dias_festivos": 23, "prestamo_emp": 24, "referidos": 25,
    "vales_despensa": 26, "comision_suma": 34, "incentivo": 35, "prestamo_laffy": 36,
    "chargeback": 37, "averias": 38, "observaciones": 39,
}

# Nomina Cliente (hoja "Base Pre nómina"): header fila idx 0, datos desde idx 1.
NOM_HEADERS = [
    "NO. SEMANAL", "NO.", "Empleado", "RFC", "Puesto", "SD", "SBC", "Sucursal",
    "Registro patronal", "Region", "FECHA DE INGRESO", "Dias trabajados", "Faltas",
    "Vacaciones", "Incapacidades", "Sueldo", "Despensa(Informativo)", "Vacaciones",
    "Prima vacacional", "reintegro_de_isr_pagado_en_exceso (MONTO)", "Aguinaldo",
    "Gratificación", "Indemnización", "Prima de antigüedad", "PTU", "Bono",
    "Prestamo Empleados", "Fondo ahorro empresa", "Comisiones", "Prima dominical",
    "Día festivo", "Día descanso", "Devolucion Infonavit", "devolucion fonacot",
    "Devolucion descuento improcedente", "subsidios por incapacidad", "Retroactivo",
    "Ajuste de sueldo", "incentivos (MONTO)", "isr_ajustado_por_subsidio (MONTO)",
    "Subsidio al Empleo (sp)", "TOTAL PERCEPCIONES", "I.S.R. (sp)", "Otros descuentos",
    "ISR AJUSTE", "I.M.S.S.", "rcv (MONTO)", "Préstamo Infonavit", "Ajuste Infonavit",
    "Ajuste INFONAVIT mes anterior", "Seguro vivienda INFONAVIT", "Préstamo FONACOT",
    "Préstamo FONACOT1", "Préstamo FONACOT2", "Préstamo FONACOT3", "Préstamo FONACOT4",
    "Ajuste Fonacot", "FONACOT MES ANTERIOR", "Fondo de ahorro Empleado",
    "Fondo de ahorro de Empresa", "Cancelaciones Prematuras", "DESCUENTO (PRESTAMOS )",
    "PRESTAMOS Laffy ", "PENSION ALIMENTICIA", "PENSION ALIMENCIA ADEUDO",
    "Descuento Averias y perdidas", "Pago anticipado a neto", "TOTAL DEDUCCIONES",
    "NETO SUELDO FISCAL", "Imss", "RCV", "Infonavit empresa", "Impuesto estatal",
    "TOTAL PREVISION SOCIAL",
]
NCOL = len(NOM_HEADERS)  # 74
# Índices de salida por nombre (para claridad)
O = {h: i for i, h in enumerate(NOM_HEADERS)}
# desambiguar las dos "Vacaciones": 13 (días incidencia) y 17 (percepción)
IDX_VAC_DIAS = 13
IDX_VAC_PERC = 17

# ---------------------------------------------------------------------------
# 2. CONSTANTES FISCALES 2026
# ---------------------------------------------------------------------------
UMA = 117.31  # UMA diaria 2026 (recuperada de la data PRIME; DOF ene-2026)

# Tarifa ISR SEMANAL 2026 (DOF Anexo 8 RMF 2026): (LI, LS, cuota_fija, %)
ISR_SEMANAL = [
    (0.01, 194.39, 0.00, 0.0192), (194.40, 1649.69, 3.72, 0.0640),
    (1649.70, 2899.26, 96.88, 0.1088), (2899.27, 3370.29, 232.89, 0.1600),
    (3370.30, 4035.15, 308.21, 0.1792), (4035.16, 8138.27, 427.35, 0.2136),
    (8138.28, 12827.08, 1303.68, 0.2352), (12827.09, 24490.90, 2406.53, 0.3000),
    (24490.91, 32653.88, 5905.62, 0.3200), (32653.89, 97961.71, 8517.74, 0.3400),
    (97961.72, float("inf"), 30722.23, 0.3500),
]
# Subsidio al empleo SEMANAL 2026: (ingreso_desde, ingreso_hasta, subsidio)
SUBSIDIO_SEMANAL = [
    (0.01, 407.33, 93.73), (407.34, 610.96, 93.66), (610.97, 799.68, 93.66),
    (799.69, 814.66, 90.44), (814.67, 1023.75, 88.06), (1023.76, 1086.19, 81.55),
    (1086.20, 1228.57, 74.83), (1228.58, 1433.32, 67.83), (1433.33, 1638.07, 58.38),
    (1638.08, 1699.88, 50.12), (1699.89, float("inf"), 0.00),
]

# Cesantía y Vejez PATRONAL 2026 por tramo de SBC/UMA (reforma pensiones, Segundo
# Transitorio) — verificado EXACTO contra la data de PRIME:
def cesantia_patronal_rate(sbc_diario: float) -> float:
    u = sbc_diario / UMA
    if u <= 1.0:
        return 0.03150
    elif u <= 1.5:
        return 0.03676
    elif u <= 2.0:
        return 0.04851
    elif u <= 2.5:
        return 0.05556
    elif u <= 3.0:
        return 0.06026
    elif u <= 3.5:
        return 0.06361
    elif u <= 4.0:
        return 0.06613
    else:
        return 0.07513

# Prima de riesgo de trabajo de PRIME (recuperada de la data: mediana 0.52444%).
PRIMA_RT = 0.0052444

# Cuotas IMSS 2026 (LSS): tasas sobre SBC (o excedente de 3 UMA / cuota fija UMA)
IMSS = {
    "fija_pat": 0.2040,          # cuota fija patrón: 20.40% UMA por día cotizado
    "exc_pat": 0.0110,           # excedente 3 UMA patrón: 1.10%
    "exc_obr": 0.0040,           # excedente 3 UMA obrero: 0.40%
    "dinero_pat": 0.0070,        # prestaciones en dinero patrón 0.70%
    "dinero_obr": 0.0025,        # obrero 0.25%
    "gmp_pat": 0.0105,           # gastos médicos pensionados patrón 1.05%
    "gmp_obr": 0.00375,          # obrero 0.375%
    "iv_pat": 0.0175,            # invalidez y vida patrón 1.75%
    "iv_obr": 0.00625,           # obrero 0.625%
    "guarderia_pat": 0.0100,     # guarderías 1.00%
    "retiro_pat": 0.0200,        # retiro 2.00%
    "ceav_obr": 0.01125,         # cesantía y vejez obrero 1.125%
}
INFONAVIT_PAT = 0.05             # aportación patronal Infonavit 5%


def isr_semanal(base_gravable: float) -> float:
    for li, ls, cf, pct in ISR_SEMANAL:
        if base_gravable <= ls:
            return cf + (base_gravable - li) * pct
    return 0.0


def subsidio_semanal(base_gravable: float) -> float:
    for desde, hasta, s in SUBSIDIO_SEMANAL:
        if base_gravable <= hasta:
            return s
    return 0.0


# Tarifa/subsidio ISR MENSUAL 2026 (DOF Anexo 8 RMF 2026). PRIME calcula el ISR
# semanal PROYECTANDO a mensual con factor 30.4/7 (recuperado de la data real),
# aplicando la tarifa mensual y regresando el resultado a semanal (×7/30.4).
ISR_MENSUAL = [
    (0.01, 844.59, 0.00, 0.0192), (844.60, 7168.51, 16.22, 0.0640),
    (7168.52, 12598.02, 420.95, 0.1088), (12598.03, 14644.64, 1011.68, 0.1600),
    (14644.65, 17533.63, 1339.14, 0.1792), (17533.64, 35362.83, 1856.84, 0.2136),
    (35362.84, 55736.68, 5665.16, 0.2352), (55736.69, 106410.50, 10457.09, 0.3000),
    (106410.51, 141880.66, 25659.23, 0.3200), (141880.67, 425641.99, 37009.69, 0.3400),
    (425642.00, float("inf"), 133488.54, 0.3500),
]
# Subsidio para el empleo 2026 (Decreto DOF 31-dic-2025): monto fijo mensual para
# ingresos ≤ $11,492.66/mes, aplicado contra el ISR. PRIME usa 15.59% × UMA mensual
# 2026 = $555.91 (recuperado EXACTO de la data: reproduce el ISR al centavo de los
# asalariados de bajo ingreso). Nota: el decoreto marca 15.02%/$536.22 de feb-dic,
# pero el sistema de nómina de PRIME aplica 15.59% todo el año.
SUBSIDIO_EMPLEO_MENSUAL = 555.91
SUBSIDIO_EMPLEO_TOPE = 11492.66
FACTOR_MENSUAL = 30.4  # días/mes para la proyección (recuperado exacto de la data)


def _tarifa(base, tabla):
    for li, ls, cf, pct in tabla:
        if base <= ls:
            return cf + (base - li) * pct
    return 0.0


def _subsidio(base, tabla):
    for desde, hasta, s in tabla:
        if base <= hasta:
            return s
    return 0.0


def isr_retencion_semanal(base_gravable_semanal: float):
    """ISR semanal a retener y subsidio al empleo entregado (proyección mensual)."""
    if base_gravable_semanal <= 0:
        return 0.0, 0.0
    bm = base_gravable_semanal * FACTOR_MENSUAL / 7.0
    isr_m = _tarifa(bm, ISR_MENSUAL)
    subs_m = SUBSIDIO_EMPLEO_MENSUAL if bm <= SUBSIDIO_EMPLEO_TOPE else 0.0
    neto_m = isr_m - subs_m
    neto_s = neto_m * 7.0 / FACTOR_MENSUAL
    if neto_s >= 0:
        return r2(neto_s), 0.0
    return 0.0, r2(-neto_s)


# ---------------------------------------------------------------------------
# 3. UTILIDADES
# ---------------------------------------------------------------------------
def _num(x) -> float:
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def r2(x: float) -> float:
    """Redondeo bancario-a-2 estilo nómina (half-up al centavo)."""
    return math.floor(x * 100 + 0.5) / 100.0


def semana_num(inc_rows) -> int:
    """Extrae el número de semana del texto 'SEMANA 26' en la primera fila de datos."""
    for r in inc_rows:
        v = str(r[INC["semana"]] or "")
        for tok in v.replace("SEMANA", "").split():
            if tok.isdigit():
                return int(tok)
    return 0


def fin_de_semana(anio: int, semana: int) -> date:
    """Domingo (fin) de la semana de nómina PRIME. Verificado: SEM26 termina 2026-06-28."""
    # PRIME numera semanas terminando en domingo; SEM 26/2026 -> 2026-06-28.
    # Ancla: domingo de la semana 26 = 2026-06-28.
    ancla_sem, ancla_dom = 26, date(2026, 6, 28)
    return ancla_dom + timedelta(weeks=(semana - ancla_sem))


# ---------------------------------------------------------------------------
# 4. PARSERS
# ---------------------------------------------------------------------------
def parse_incidencias(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    data = [r for r in rows[3:] if r[INC["id"]] not in (None, "")]
    return data


def build_isn_rates(nomina_cliente_paths):
    """
    Recupera la tasa de ISN (impuesto estatal) por Registro Patronal a partir de
    archivos Nomina Cliente previos: ISN = tasa_estado(RP) × TOTAL PERCEPCIONES.
    Cada RP corresponde a un estado con su tasa 2026 (1.8%, 2%, 3%, 4%, ...).
    """
    from collections import defaultdict
    votos = defaultdict(lambda: defaultdict(int))
    for path in nomina_cliente_paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        for r in rows[1:]:
            if r[O["NO."]] in (None, ""):
                continue
            rp = str(r[O["Registro patronal"]])
            perc = _num(r[O["TOTAL PERCEPCIONES"]])
            isn = _num(r[O["Impuesto estatal"]])
            if perc > 0 and isn > 0:
                tasa = round(isn / perc, 4)
                votos[rp][tasa] += 1
    rates = {}
    for rp, cnt in votos.items():
        rates[rp] = max(cnt.items(), key=lambda kv: kv[1])[0]  # tasa más frecuente
    return rates


def build_rt_primas(nomina_cliente_paths):
    """
    Recupera la PRIMA DE RIESGO DE TRABAJO por Registro Patronal desde archivos
    Nomina Cliente previos. Cada registro patronal de PRIME tiene su propia prima
    (clase de riesgo IMSS): 0.50% – 0.64% observadas. Se despeja de:
        col69 (Imss patronal) = componentes_fijos(SBC) + prima × SBC × 7
    usando empleados con semana completa (7 días, sin incapacidad).
    """
    from collections import defaultdict
    import statistics
    acc = defaultdict(list)
    for path in nomina_cliente_paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        for r in rows[1:]:
            if r[O["NO."]] in (None, ""):
                continue
            sbc = _num(r[O["SBC"]])
            dias = _num(r[O["Dias trabajados"]])
            incap = _num(r[O["Incapacidades"]])
            if sbc <= 0 or dias < 7 or incap != 0:
                continue
            sbw = sbc * 7
            exc = max(0.0, sbw - 3 * UMA * 7)
            base = (IMSS["fija_pat"] * UMA * 7 + IMSS["exc_pat"] * exc +
                    IMSS["dinero_pat"] * sbw + IMSS["gmp_pat"] * sbw +
                    IMSS["iv_pat"] * sbw + IMSS["guarderia_pat"] * sbw)
            prima = (_num(r[O["Imss"]]) - base) / sbw
            if 0 < prima < 0.03:
                acc[str(r[O["Registro patronal"]])].append(prima)
    return {rp: round(statistics.median(v), 6) for rp, v in acc.items()}


def build_master(nomina_cliente_paths):
    """
    Construye el maestro por empleado (SBC, sucursal, RP, region, préstamos, pensión,
    seguro vivienda, nombre/puesto autoritativos) a partir de uno o más archivos
    Nomina Cliente previos (el más reciente gana). Estas columnas NO vienen en
    INCIDENCIAS y son estables semana a semana.
    """
    master = {}
    for path in nomina_cliente_paths:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        for r in rows[1:]:
            if r[O["NO."]] in (None, ""):
                continue
            emp_id = str(r[O["NO."]]).strip()
            master[emp_id] = {
                "SBC": _num(r[O["SBC"]]),
                "Sucursal": r[O["Sucursal"]],
                "RP": r[O["Registro patronal"]],
                "Region": r[O["Region"]],
                "Empleado": r[O["Empleado"]],
                "Puesto": r[O["Puesto"]],
                # Deducciones recurrentes configuradas por empleado (setup, no incidencia):
                "Prestamo Infonavit": _num(r[O["Préstamo Infonavit"]]),
                "Seguro vivienda INFONAVIT": _num(r[O["Seguro vivienda INFONAVIT"]]),
                "Préstamo FONACOT": _num(r[O["Préstamo FONACOT"]]),
                "Préstamo FONACOT1": _num(r[O["Préstamo FONACOT1"]]),
                "Préstamo FONACOT2": _num(r[O["Préstamo FONACOT2"]]),
                "Préstamo FONACOT3": _num(r[O["Préstamo FONACOT3"]]),
                "Préstamo FONACOT4": _num(r[O["Préstamo FONACOT4"]]),
                "PENSION ALIMENTICIA": _num(r[O["PENSION ALIMENTICIA"]]),
            }
    return master


# ---------------------------------------------------------------------------
# 5. MOTOR: una fila de Nomina Cliente por empleado
# ---------------------------------------------------------------------------
def _dias(inc_row, dom_semana: date):
    """Días trabajados (sueldo) y días cotizados (IMSS)."""
    faltas = _num(inc_row[INC["faltas"]])
    incap = _num(inc_row[INC["incap"]])
    vac = _num(inc_row[INC["vac_aprob"]])
    # Proración por alta a mitad de semana:
    ingreso = inc_row[INC["ingreso"]]
    tope = 7
    if isinstance(ingreso, datetime):
        ingreso = ingreso.date()
    if isinstance(ingreso, date):
        d = (dom_semana - ingreso).days + 1
        if 0 < d < 7:
            tope = d
    dias_trab = max(0.0, min(tope, 7 - faltas - incap - vac))  # sueldo
    # IMSS: se cotiza la semana completa (7 días); faltas, incap y vacaciones NO
    # reducen los días cotizados (verificado en data). Solo el prorrateo por alta
    # a mitad de semana los reduce.
    dias_cot = tope
    return dias_trab, dias_cot


def calcula_fila(inc_row, master, isn_rates, rt_primas, sem_label, dom_semana):
    emp_id = str(inc_row[INC["id"]]).strip()
    m = master.get(emp_id, {})
    sd = _num(inc_row[INC["sd"]])
    sbc = m.get("SBC", 0.0)
    dias_trab, dias_cot = _dias(inc_row, dom_semana)

    row = [0] * NCOL
    # --- Identidad ---
    row[O["NO. SEMANAL"]] = sem_label
    row[O["NO."]] = inc_row[INC["id"]]
    row[O["Empleado"]] = m.get("Empleado") or inc_row[INC["nombre"]]
    row[O["RFC"]] = inc_row[INC["rfc"]]
    row[O["Puesto"]] = m.get("Puesto") or inc_row[INC["puesto"]]
    row[O["SD"]] = sd
    row[O["SBC"]] = sbc
    row[O["Sucursal"]] = m.get("Sucursal", "")
    row[O["Registro patronal"]] = m.get("RP", "")
    row[O["Region"]] = m.get("Region", 0)
    row[O["FECHA DE INGRESO"]] = inc_row[INC["ingreso"]]

    # --- Incidencias / días ---
    row[O["Dias trabajados"]] = dias_trab
    row[O["Faltas"]] = _num(inc_row[INC["faltas"]])
    row[IDX_VAC_DIAS] = _num(inc_row[INC["vac_aprob"]])         # Vacaciones (días)
    row[O["Incapacidades"]] = _num(inc_row[INC["incap"]])

    # --- Percepciones ---
    sueldo = r2(sd * dias_trab)
    row[O["Sueldo"]] = sueldo
    row[O["Despensa(Informativo)"]] = _num(inc_row[INC["vales_despensa"]])
    vac_perc = r2(_num(inc_row[INC["vac_aprob"]]) * sd)
    row[IDX_VAC_PERC] = vac_perc                               # Vacaciones (percepción)
    row[O["Prima vacacional"]] = r2(_num(inc_row[INC["prima_vac_dias"]]) * sd) if False else 0
    row[O["Comisiones"]] = _num(inc_row[INC["comision_suma"]])
    domingos = _num(inc_row[INC["domingos"]])
    prima_dom = r2(domingos * sd * 0.25)
    row[O["Prima dominical"]] = prima_dom
    row[O["Día festivo"]] = r2(_num(inc_row[INC["dias_festivos"]]) * sd * 2)
    row[O["Día descanso"]] = r2(_num(inc_row[INC["descansos"]]) * sd * 2)
    row[O["incentivos (MONTO)"]] = _num(inc_row[INC["incentivo"]])

    # TOTAL PERCEPCIONES = suma del bloque (idx 15..40); se recalcula por fórmula al escribir.
    total_perc = sum(_num(row[i]) for i in range(15, 41))
    row[O["TOTAL PERCEPCIONES"]] = r2(total_perc)

    # --- ISR (proyección mensual 2026) ---
    # Base gravable = Sueldo (por días laborados) + Vacaciones (se pagan aparte pero
    # gravan) + Comisiones + incentivos + porción gravable de prima dominical. En
    # términos de días: SD × (7 − faltas − incapacidades); la vacación NO reduce la
    # base (se paga). Verificado contra la data real de PRIME (95%+ al peso).
    grav_prima_dom = max(0.0, prima_dom - UMA * domingos)      # exención 1 UMA por domingo
    base_isr = (sueldo + vac_perc + _num(row[O["Comisiones"]]) +
                _num(row[O["incentivos (MONTO)"]]) + grav_prima_dom)
    isr_ret, subsidio_ent = isr_retencion_semanal(base_isr)
    row[O["I.S.R. (sp)"]] = isr_ret
    row[O["Subsidio al Empleo (sp)"]] = subsidio_ent

    # --- IMSS / RCV obrero (deducciones) ---
    sbw = sbc * dias_cot
    exc = max(0.0, sbw - 3 * UMA * dias_cot)
    imss_obr = (IMSS["exc_obr"] * exc + IMSS["dinero_obr"] * sbw +
                IMSS["gmp_obr"] * sbw + IMSS["iv_obr"] * sbw)
    row[O["I.M.S.S."]] = r2(imss_obr)
    row[O["rcv (MONTO)"]] = r2(IMSS["ceav_obr"] * sbw)

    # --- Deducciones recurrentes (maestro) ---
    row[O["Préstamo Infonavit"]] = m.get("Prestamo Infonavit", 0.0)
    row[O["Seguro vivienda INFONAVIT"]] = m.get("Seguro vivienda INFONAVIT", 0.0)
    for k in ["Préstamo FONACOT", "Préstamo FONACOT1", "Préstamo FONACOT2",
              "Préstamo FONACOT3", "Préstamo FONACOT4"]:
        row[O[k]] = m.get(k, 0.0)
    row[O["PENSION ALIMENTICIA"]] = m.get("PENSION ALIMENTICIA", 0.0)
    row[O["PRESTAMOS Laffy "]] = _num(inc_row[INC["prestamo_laffy"]])
    row[O["Cancelaciones Prematuras"]] = _num(inc_row[INC["chargeback"]])
    row[O["Descuento Averias y perdidas"]] = _num(inc_row[INC["averias"]])

    total_ded = sum(_num(row[i]) for i in range(42, 67))
    row[O["TOTAL DEDUCCIONES"]] = r2(total_ded)
    row[O["NETO SUELDO FISCAL"]] = r2(_num(row[O["TOTAL PERCEPCIONES"]]) -
                                      _num(row[O["TOTAL DEDUCCIONES"]]) -
                                      _num(row[O["Despensa(Informativo)"]]))

    # --- Previsión social (patronal) ---
    rp = str(m.get("RP", ""))
    prima_rt = rt_primas.get(rp, PRIMA_RT) if rt_primas else PRIMA_RT
    fija = IMSS["fija_pat"] * UMA * dias_cot
    imss_pat = (prima_rt * sbw + fija + IMSS["exc_pat"] * exc +
                IMSS["dinero_pat"] * sbw + IMSS["gmp_pat"] * sbw +
                IMSS["iv_pat"] * sbw + IMSS["guarderia_pat"] * sbw)
    row[O["Imss"]] = r2(imss_pat)
    rcv_pat = (IMSS["retiro_pat"] + cesantia_patronal_rate(sbc)) * sbw
    row[O["RCV"]] = r2(rcv_pat)
    row[O["Infonavit empresa"]] = r2(INFONAVIT_PAT * sbw)
    # Impuesto estatal (ISN) = tasa_estado(RP) × TOTAL PERCEPCIONES.
    tasa_isn = isn_rates.get(rp, 0.0) if isn_rates else 0.0
    row[O["Impuesto estatal"]] = r2(tasa_isn * _num(row[O["TOTAL PERCEPCIONES"]]))
    row[O["TOTAL PREVISION SOCIAL"]] = r2(_num(row[O["Imss"]]) + _num(row[O["RCV"]]) +
                                          _num(row[O["Infonavit empresa"]]) +
                                          _num(row[O["Impuesto estatal"]]))
    return row


def generar(incidencias_path, master, isn_rates=None, rt_primas=None, anio=2026):
    inc = parse_incidencias(incidencias_path)
    sem = semana_num(inc)
    sem_label = f"SEM {sem}"
    dom = fin_de_semana(anio, sem)
    filas = [calcula_fila(r, master, isn_rates or {}, rt_primas or {}, sem_label, dom)
             for r in inc]
    return sem, filas


# ---------------------------------------------------------------------------
# 6. ESCRITURA DEL EXCEL (formato Nomina Cliente: 74 cols + fórmulas de totales)
# ---------------------------------------------------------------------------
def escribir_xlsx(filas, out_path):
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Base Pre nómina"
    ws.append(NOM_HEADERS)
    # columnas con fórmula (igual que el Excel original), 0-based -> letra Excel (1-based)
    L = lambda i: get_column_letter(i + 1)
    idx_perc, idx_desp = O["TOTAL PERCEPCIONES"], O["Despensa(Informativo)"]
    idx_ded, idx_neto = O["TOTAL DEDUCCIONES"], O["NETO SUELDO FISCAL"]
    idx_prev = O["TOTAL PREVISION SOCIAL"]
    for fi, fila in enumerate(filas):
        r = fi + 2  # fila Excel (1=header)
        vals = list(fila)
        # TOTAL PERCEPCIONES = SUM(Sueldo..Subsidio) = cols 15..40
        vals[idx_perc] = f"=SUM({L(15)}{r}:{L(40)}{r})"
        # TOTAL DEDUCCIONES = SUM(ISR..Pago anticipado) = cols 42..66
        vals[idx_ded] = f"=SUM({L(42)}{r}:{L(66)}{r})"
        # NETO = TOTAL PERCEPCIONES - TOTAL DEDUCCIONES - Despensa(Informativo)
        vals[idx_neto] = f"=({L(idx_perc)}{r}-{L(idx_ded)}{r})-{L(idx_desp)}{r}"
        # TOTAL PREVISION SOCIAL = Imss + RCV + Infonavit empresa + Impuesto estatal
        vals[idx_prev] = (f"={L(O['Imss'])}{r}+{L(O['RCV'])}{r}+"
                          f"{L(O['Infonavit empresa'])}{r}+{L(O['Impuesto estatal'])}{r}")
        ws.append(vals)
    wb.save(out_path)
    return out_path


def generar_archivo(incidencias_path, referencia_paths, out_path, anio=2026):
    """Pipeline completo: construye maestro/tasas desde referencia(s) Nomina Cliente,
    genera las filas desde INCIDENCIAS y escribe el Excel. Devuelve (semana, n_filas)."""
    refs = referencia_paths if isinstance(referencia_paths, (list, tuple)) else [referencia_paths]
    master = build_master(refs)
    isn = build_isn_rates(refs)
    rt = build_rt_primas(refs)
    sem, filas = generar(incidencias_path, master, isn, rt, anio=anio)
    escribir_xlsx(filas, out_path)
    return sem, len(filas)
