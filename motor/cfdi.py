"""Parser del CFDI One Facture (insumo 3) — spec §5.3.5 + reglas v6.

Determinista, centavos enteros. Regla de neto CFDI (v6 §5.3.5, CONFIRMADA):
    neto_cfdi = Total percepciones + Total otros pagos - Total deducciones
NUNCA usar la columna "Total" para conciliar percep/deduc.

- Excluye timbres duplicados (por UUID; fallback RFC+fecha+total) de las sumas.
- Excluye CFDI cancelados.
- Filtra por mes de pago (fecha pago) para acotar al periodo del informe.
- Agrega por RFC (llave de conciliación a nivel empleado) y total.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from openpyxl import load_workbook

from .normaliza import a_centavos, a_iso, normalizar_rfc, normalizar_clave

# posiciones LEGACY (enero-2026, 81 cols) — solo fallback. One Facture agrega
# columnas y corre las posiciones (abril-2026 = 92 cols: los totales pasan de
# 72/75/77/78 a 83/86/88/89), así que resolvemos por NOMBRE contra el encabezado.
COL = {
    "periodo": 0, "uuid": 4, "tipo_nomina": 7, "fecha_pago": 12,
    "registro_patronal": 21, "rfc": 22, "estado": 43,
    "estado_cancelacion": 45, "estado_proceso_cancelacion": 46,
    "total": 72, "total_deducciones": 75, "total_otros_pagos": 77,
    "total_percepciones": 78,
}
# nombre exacto en el encabezado (match por clave normalizada) por cada columna
NOMBRES_CFDI = {
    "periodo": "Periodo", "uuid": "UUID", "tipo_nomina": "Tipo nomina",
    "fecha_pago": "Fecha pago", "registro_patronal": "Registro patronal",
    "rfc": "RFC receptor", "estado": "Estado",
    "estado_cancelacion": "Estado cancelacion",
    "estado_proceso_cancelacion": "Estado proceso cancelacion",
    "total": "Total", "total_deducciones": "Total deducciones",
    "total_otros_pagos": "Total otros pagos",
    "total_percepciones": "Total percepciones",
}
HOJA_CFDI = "Datos Cfdis"


def _resolver_columnas(header):
    """Mapea clave interna -> índice por NOMBRE del encabezado; si un nombre no
    aparece, cae a la posición legacy (enero). Robusto a columnas añadidas."""
    norm_to_idx = {}
    for j, c in enumerate(header):
        if isinstance(c, str) and c.strip():
            norm_to_idx.setdefault(normalizar_clave(c), j)
    return {clave: norm_to_idx.get(normalizar_clave(nombre), COL.get(clave))
            for clave, nombre in NOMBRES_CFDI.items()}


@dataclass
class AgregadoRFC:
    cfdis: int = 0
    percepciones: int = 0
    otros_pagos: int = 0
    deducciones: int = 0
    neto: int = 0            # perc + otros - ded
    total_col: int = 0       # columna "Total" (referencia, NO se usa para conciliar)
    uuids: list = field(default_factory=list)


@dataclass
class ResultadoCFDI:
    por_rfc: dict            # rfc -> AgregadoRFC
    total: AgregadoRFC
    meta: dict


def _es_cancelado(fila, col):
    for k in ("estado", "estado_cancelacion", "estado_proceso_cancelacion"):
        j = col.get(k)
        if j is None or j >= len(fila):
            continue
        v = fila[j]
        if v and normalizar_clave(v) in ("CANCELADO", "CANCELADA"):
            return True
    return False


def parse_cfdi(path, mes_pago=None):
    """mes_pago: 'YYYY-MM' para filtrar por fecha de pago; None = todo el archivo."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[HOJA_CFDI] if HOJA_CFDI in wb.sheetnames else \
        max(wb.worksheets, key=lambda s: (s.max_row or 0) * (s.max_column or 0))
    it = ws.iter_rows(values_only=True)
    header = next(it)                       # encabezado (fila 0)
    col = _resolver_columnas(header)        # índices por nombre (enero y abril)

    por_rfc: dict = defaultdict(AgregadoRFC)
    total = AgregadoRFC()
    vistos_uuid = set()
    vistos_rfc_fecha_total = set()
    n = dup = canc = fuera_mes = sin_rfc = 0

    for fila in it:
        if fila is None or not any(v is not None and str(v).strip() for v in fila):
            continue
        n += 1

        # filtro por mes de pago
        fpago = a_iso(fila[col["fecha_pago"]]) if fila[col["fecha_pago"]] not in (None, "") else None
        if mes_pago and (not fpago or not fpago.startswith(mes_pago)):
            fuera_mes += 1
            continue

        if _es_cancelado(fila, col):
            canc += 1
            continue

        rfc, _ = normalizar_rfc(fila[col["rfc"]] if col["rfc"] < len(fila) else "")
        perc = a_centavos(fila[col["total_percepciones"]])
        otros = a_centavos(fila[col["total_otros_pagos"]])
        ded = a_centavos(fila[col["total_deducciones"]])
        tot = a_centavos(fila[col["total"]])
        neto = perc + otros - ded

        # dedup: UUID primario; fallback (rfc, fecha, total)
        uuid = fila[col["uuid"]]
        if uuid:
            if uuid in vistos_uuid:
                dup += 1
                continue
            vistos_uuid.add(uuid)
        else:
            k = (rfc, fpago, tot)
            if k in vistos_rfc_fecha_total:
                dup += 1
                continue
            vistos_rfc_fecha_total.add(k)

        if not rfc:
            sin_rfc += 1

        for bucket in (por_rfc[rfc], total):
            bucket.cfdis += 1
            bucket.percepciones += perc
            bucket.otros_pagos += otros
            bucket.deducciones += ded
            bucket.neto += neto
            bucket.total_col += tot
        if uuid:
            por_rfc[rfc].uuids.append(uuid)

    wb.close()
    meta = {
        "archivo": path, "hoja": ws.title, "mes_pago": mes_pago,
        "filas_totales": n, "cfdis_contados": total.cfdis,
        "duplicados_excluidos": dup, "cancelados_excluidos": canc,
        "fuera_de_mes": fuera_mes, "sin_rfc": sin_rfc,
        "rfcs_unicos": len(por_rfc),
    }
    return ResultadoCFDI(dict(por_rfc), total, meta)
