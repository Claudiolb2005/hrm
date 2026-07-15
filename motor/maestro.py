"""Parser del MAESTRO (insumo 2) — hojas RESUMEN y BASES (spec §5.3.4).

- RESUMEN: ISN por estado ya calculado (RP, entidad, empleados, base, tasa, impuesto)
  → alimenta la Sección X. Validación: base*tasa ≈ impuesto por renglón.
- BASES: mapeo registro patronal → entidad federativa + tasas + Vales/PTU exento
  → construye el mapeo estado↔RP que usan las Secciones VI y IX.
"""
from __future__ import annotations
from openpyxl import load_workbook
from .normaliza import a_centavos, normalizar_clave


def _filas(path, hoja):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[hoja]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    return filas


def _fila_encabezado(filas, etiquetas):
    """Devuelve idx de la fila que contiene todas las etiquetas dadas (normalizadas)."""
    objetivo = {normalizar_clave(e) for e in etiquetas}
    for i, fila in enumerate(filas[:6]):
        presentes = {normalizar_clave(c) for c in fila if isinstance(c, str)}
        if objetivo <= presentes:
            return i
    return 0


def _col_idx(header, etiqueta):
    n = normalizar_clave(etiqueta)
    for j, c in enumerate(header):
        if isinstance(c, str) and normalizar_clave(c) == n:
            return j
    return None


def parse_resumen(path, hoja="RESUMEN"):
    """→ lista de dicts {rp, entidad, empleados, base_isn, tasa, impuesto, cuadra}."""
    filas = _filas(path, hoja)
    h = _fila_encabezado(filas, ["REGISTRO PATRONAL", "BASE ISN", "IMPUESTO"])
    header = filas[h]
    ci = {k: _col_idx(header, v) for k, v in {
        "rp": "REGISTRO PATRONAL", "entidad": "ENTIDAD FEDERATIVA",
        "empleados": "No. TRABAJADORES EMISION", "base_isn": "BASE ISN",
        "tasa": "TASA", "impuesto": "IMPUESTO",
    }.items()}
    out = []
    for fila in filas[h + 1:]:
        if not fila or ci["rp"] is None:
            continue
        rp = fila[ci["rp"]]
        if not rp or normalizar_clave(rp) in ("", "TOTAL", "TOTALES", "SUMA"):
            continue
        base = a_centavos(fila[ci["base_isn"]]) if ci["base_isn"] is not None else 0
        imp = a_centavos(fila[ci["impuesto"]]) if ci["impuesto"] is not None else 0
        tasa = fila[ci["tasa"]] if ci["tasa"] is not None else None
        try:
            tasa_f = float(tasa) if tasa not in (None, "") else 0.0
        except (TypeError, ValueError):
            tasa_f = 0.0
        emp = fila[ci["empleados"]] if ci["empleados"] is not None else None
        # validación base*tasa ≈ impuesto (±1 peso por redondeo de tasa)
        cuadra = abs(round(base * tasa_f) - imp) <= 100
        out.append({
            "rp": str(rp).strip(),
            "entidad": str(fila[ci["entidad"]]).strip() if ci["entidad"] is not None and fila[ci["entidad"]] else "",
            "empleados": int(emp) if isinstance(emp, (int, float)) else None,
            "base_isn": base, "tasa": tasa_f, "impuesto": imp, "cuadra": cuadra,
        })
    return out


def parse_bases(path, hoja="BASES"):
    """→ dict rp -> {entidad, tasa, vales_gravado, ptu_gravado}."""
    filas = _filas(path, hoja)
    # el encabezado solo garantiza RP + ENTIDAD; la columna de tasa cambia de
    # nombre entre entregas ("TASA" en enero-2026, "ISN 2026" en abril-2026).
    h = _fila_encabezado(filas, ["REGISTRO PATRONAL", "ENTIDAD FEDERATIVA"])
    header = filas[h]

    def col(*etiquetas):
        for e in etiquetas:
            j = _col_idx(header, e)
            if j is not None:
                return j
        return None

    ci = {
        "rp": col("REGISTRO PATRONAL"), "entidad": col("ENTIDAD FEDERATIVA"),
        "tasa": col("TASA", "ISN 2026"), "vales": col("VALES"), "ptu": col("PTU"),
    }
    out = {}
    for fila in filas[h + 1:]:
        if not fila or ci["rp"] is None or not fila[ci["rp"]]:
            continue
        rp = str(fila[ci["rp"]]).strip()
        if normalizar_clave(rp) in ("", "TOTAL", "TOTALES"):
            continue
        def grav(col):
            v = fila[ci[col]] if ci[col] is not None else None
            return normalizar_clave(v) == "GRAVADO" if v is not None else False
        out[rp] = {
            "entidad": str(fila[ci["entidad"]]).strip() if ci["entidad"] is not None and fila[ci["entidad"]] else "",
            "tasa": float(fila[ci["tasa"]]) if ci["tasa"] is not None and isinstance(fila[ci["tasa"]], (int, float)) else 0.0,
            "vales_gravado": grav("vales"), "ptu_gravado": grav("ptu"),
        }
    return out
